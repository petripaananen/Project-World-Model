"""
Populates the official 'Build with Gemini XPRIZE - PL Template.xlsx' spreadsheet
and generates the exact corresponding 1-Page PDF:
'Docs/Build_with_Gemini_XPRIZE_PL_Statement.pdf'
"""

import os
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOCS_DIR = os.path.join(BASE_DIR, "Docs")
EXCEL_PATH = os.path.join(DOCS_DIR, "Build with Gemini XPRIZE - PL Template.xlsx")
POPULATED_EXCEL = os.path.join(DOCS_DIR, "Project_World_Model_XPRIZE_PL_Statement.xlsx")
OUTPUT_PDF = os.path.join(DOCS_DIR, "Build_with_Gemini_XPRIZE_PL_Statement.pdf")

def populate_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Template']

    # Project Name / Description header
    ws['B4'] = "Project: Project World Model (PWM) | Track: Small Business Services"
    ws['B6'] = "Founder: Petri Paananen | Status: Pre-Launch / Closed Beta ($0 Revenue)"

    # Revenue rows (Row 9: Independent Sales, Row 10: Related Party Revenue)
    # Columns C: May, D: June, E: July, F: August
    ws['C9'] = 0.00
    ws['D9'] = 0.00
    ws['E9'] = 0.00
    ws['F9'] = 0.00

    ws['C10'] = 0.00
    ws['D10'] = 0.00
    ws['E10'] = 0.00
    ws['F10'] = 0.00

    # Expenses: COGS
    # Row 15: Personnel
    ws['C15'] = 0.00
    ws['D15'] = 0.00
    ws['E15'] = 0.00
    ws['F15'] = 0.00

    # Row 16: Software Subscriptions (Artifact Registry & GCP storage)
    ws['C16'] = 0.00
    ws['D16'] = 0.00
    ws['E16'] = 2.00   # ~$2.00 USD (€1.83)
    ws['F16'] = 0.00

    # Row 17: Tokens (Gemini API Tokens)
    ws['C17'] = 0.00
    ws['D17'] = 0.00
    ws['E17'] = 36.00  # ~$36.00 USD (€33.01)
    ws['F17'] = 0.00

    # Expenses: SG&A
    # Row 19: Personnel
    ws['C19'] = 0.00
    ws['D19'] = 0.00
    ws['E19'] = 0.00
    ws['F19'] = 0.00

    # Row 20: Software Subscriptions
    ws['C20'] = 0.00
    ws['D20'] = 0.00
    ws['E20'] = 0.00
    ws['F20'] = 0.00

    # Row 21: Tokens
    ws['C21'] = 0.00
    ws['D21'] = 0.00
    ws['E21'] = 0.00
    ws['F21'] = 0.00

    # Row 23: Other Expenses
    ws['C23'] = 0.00
    ws['D23'] = 0.00
    ws['E23'] = 0.00
    ws['F23'] = 0.00

    # Save both original template and named output
    wb.save(EXCEL_PATH)
    wb.save(POPULATED_EXCEL)
    print(f"[SUCCESS] Populated Excel template saved: {EXCEL_PATH}")

def generate_pdf_from_template():
    doc = SimpleDocTemplate(
        OUTPUT_PDF,
        pagesize=letter,
        leftMargin=30,
        rightMargin=30,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    primary_color = colors.HexColor("#0f172a") # dark slate
    accent_color = colors.HexColor("#0284c7")  # blue accent
    text_color = colors.HexColor("#1e293b")    # slate text
    light_bg = colors.HexColor("#f8fafc")      # very light slate
    header_bg = colors.HexColor("#1e293b")     # table header slate
    section_bg = colors.HexColor("#e2e8f0")    # subheader row
    total_bg = colors.HexColor("#cbd5e1")      # total row highlight

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=primary_color,
        spaceAfter=2
    )

    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=accent_color,
        spaceAfter=6
    )

    body_style = ParagraphStyle(
        'BodyDark',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        leading=10,
        textColor=text_color,
        spaceAfter=2
    )

    th_style = ParagraphStyle(
        'TH',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9.5,
        textColor=colors.white,
        alignment=1
    )

    td_style = ParagraphStyle(
        'TD',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.2,
        leading=9,
        textColor=text_color,
        alignment=1
    )

    td_left = ParagraphStyle(
        'TDLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.2,
        leading=9,
        textColor=text_color,
        alignment=0
    )

    td_left_bold = ParagraphStyle(
        'TDLeftBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9.5,
        textColor=primary_color,
        alignment=0
    )

    story = []

    # Header mirroring template
    story.append(Paragraph("BUILD WITH GEMINI XPRIZE", title_style))
    story.append(Paragraph("<b>PROFIT & LOSS STATEMENT</b> | Program Period: May 19 – August 17 | Currency: <b>USD ($)</b>", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.2, color=accent_color, spaceAfter=5))

    # Meta Info Table
    meta_data = [
        [
            Paragraph("<b>Project:</b> Project World Model (PWM)<br/><b>Category:</b> Small Business Services Track", body_style),
            Paragraph("<b>Founder:</b> Petri Paananen<br/><b>Operational State:</b> Pre-Launch / Closed Beta ($0 Revenue)", body_style)
        ]
    ]
    t_meta = Table(meta_data, colWidths=[270, 282])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_bg),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#94a3b8")),
        ('PADDING', (0,0), (-1,-1), 4)
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 6))

    # P&L Template Table
    table_rows = [
        # Table Header
        [
            Paragraph("Description", th_style),
            Paragraph("May", th_style),
            Paragraph("June", th_style),
            Paragraph("July", th_style),
            Paragraph("August", th_style),
            Paragraph("Full 90 Days", th_style)
        ],
        # REVENUE Section
        [Paragraph("<b>REVENUE</b>", td_left_bold), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style)],
        [Paragraph("Independent Sales (ie. sales of product or service)", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        [Paragraph("Related Party Revenue (ie. see Rules)", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        [Paragraph("<b>TOTAL REVENUE</b>", td_left_bold), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style)],
        
        # EXPENSES Section
        [Paragraph("<b>EXPENSES</b>", td_left_bold), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style)],
        
        # COGS Sub-section
        [Paragraph("<b>&nbsp;&nbsp;COGS</b>", td_left_bold), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Personnel", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Software Subscriptions (Artifact Registry & Storage)", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$2.00", td_style), Paragraph("$0.00", td_style), Paragraph("$2.00", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Tokens (Google Vertex AI / Gemini API)", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$36.00", td_style), Paragraph("$0.00", td_style), Paragraph("$36.00", td_style)],
        
        # SG&A Sub-section
        [Paragraph("<b>&nbsp;&nbsp;SG&A</b>", td_left_bold), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Personnel", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Software Subscriptions", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Tokens", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        
        # Other Expenses
        [Paragraph("<b>&nbsp;&nbsp;Other Expenses</b>", td_left_bold), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style), Paragraph("", td_style)],
        [Paragraph("&nbsp;&nbsp;&nbsp;&nbsp;• Other expenses (see Legend)", td_left), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style), Paragraph("$0.00", td_style)],
        
        # TOTAL EXPENSES
        [Paragraph("<b>TOTAL EXPENSES</b>", td_left_bold), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$38.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$38.00</b>", td_style)],
        
        # PROFIT (LOSS)
        [Paragraph("<b>PROFIT (LOSS)</b>", td_left_bold), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>-$38.00</b>", td_style), Paragraph("<b>$0.00</b>", td_style), Paragraph("<b>-$38.00</b>", td_style)]
    ]

    t_pnl = Table(table_rows, colWidths=[242, 60, 60, 65, 60, 65])
    t_pnl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ('BACKGROUND', (0, 1), (-1, 1), section_bg),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor("#f1f5f9")),
        ('BACKGROUND', (0, 5), (-1, 5), section_bg),
        ('BACKGROUND', (0, 6), (-1, 6), light_bg),
        ('BACKGROUND', (0, 10), (-1, 10), light_bg),
        ('BACKGROUND', (0, 14), (-1, 14), light_bg),
        ('BACKGROUND', (0, 16), (-1, 16), colors.HexColor("#e2e8f0")),
        ('BACKGROUND', (0, 17), (-1, 17), total_bg),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(t_pnl)
    story.append(Spacer(1, 5))

    # Official Legend & Disclosures (matching XPRIZE template legend)
    legend_header = ParagraphStyle('LH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=primary_color)
    legend_body = ParagraphStyle('LB', parent=styles['Normal'], fontName='Helvetica', fontSize=6.8, leading=8.5, textColor=text_color)
    
    story.append(Paragraph("<b>OFFICIAL TEMPLATE LEGEND & DISCLOSURES</b>", legend_header))
    story.append(Paragraph(
        "• <b>Independent Sales:</b> Sales to arms-length third parties ($0.00 pre-launch beta).<br/>"
        "• <b>Related-Party Revenue:</b> Sales to founders, family, or affiliates ($0.00).<br/>"
        "• <b>COGS (Tokens & Subscriptions):</b> Direct costs to build and run the model on GCP (€34.92 / ~$38.00 USD covering 4.5M+ Gemini API tokens and container storage).<br/>"
        "• <b>SG&A:</b> Operating overhead, sales, marketing, and founder labor contributed via sweat-equity ($0.00).<br/>"
        "• <i>Build with Gemini XPRIZE | Managed by Devpost | CONFIDENTIAL | Certified accurate by Petri Paananen (Founder)</i>",
        legend_body
    ))

    doc.build(story)
    print(f"[SUCCESS] Official Template PDF generated: {OUTPUT_PDF}")
    print(f"File size: {os.path.getsize(OUTPUT_PDF) / 1024:.2f} KB")

if __name__ == "__main__":
    populate_excel()
    generate_pdf_from_template()
